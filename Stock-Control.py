import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
import sqlite3 as sql
from Settings import *
from openpyxl import Workbook
from tkinter import filedialog
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def create_table():
    # Create or connect to the database and create necessary tables if not exist
    vt = sql.connect("Stok.sqlite")
    imlec = vt.cursor()
    imlec.execute("CREATE TABLE IF NOT EXISTS Musteri (Adi,Soyadi,Telefon,E_posta,Adres)")
    imlec.execute("CREATE TABLE IF NOT EXISTS Urun (UrunAdi,UrunKodu,UrunSayisi,UrunMaliyet,UrunSatis)")
    imlec.execute("CREATE TABLE IF NOT EXISTS Satis (Adi,Soyadi,Telefon,UrunAdi,UrunKodu,Miktar,Tutar)")


create_table()
window = ctk.CTk()
window.title("Stok Kontrol")
window.geometry("1300x600")
window.columnconfigure((0,1,2,3,4), weight=1, uniform="a")
window.rowconfigure((0, 1, 2, 3, 4), weight=1, uniform="a")
ozet = ttk.Treeview(window,columns=("Adı","Soyadı","Telefon","Ürün Adı","Ürün Kodu","Miktar","Tutar"),show="headings")
ozet.grid(column=1,columnspan=4, row=0,rowspan=4, sticky="ns", padx=20, pady=20)
ozet.heading("Adı",text="Adı")
ozet.heading("Soyadı",text="Soyadı")
ozet.heading("Telefon",text="Telefon")
ozet.heading("Ürün Adı",text="Ürün Adı")
ozet.heading("Ürün Kodu",text="Ürün Kodu")
ozet.heading("Miktar",text="Miktar")
ozet.heading("Tutar",text="Tutar")
button_font = ctk.CTkFont(family=FONT, size=MAIN_TEXT_SIZE, weight="bold")
label = ctk.CTkLabel(window,text="Net Kâr:",text_color="Green")
label.grid(column=3, row=4,sticky="e")
text = ctk.CTkEntry(window, text_color="Green", bg_color="#242323", fg_color="#242323",border_color="#242323")
text.grid(column=4, row=4,sticky="w", padx=10, pady=40)


# Function to export summary to Excel
def export_to_excel(ozet, filename):
    wb = Workbook()
    ws = wb.active
    # Add header rows and set them to blue color
    columns = ozet["columns"]
    for col_idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Mavi renk
        # Set cell size
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(column_name) + 2, 10)  # Minimum 10 karakter genişlik
    # Transfer data from Treeview to Excel
    for row_idx, item in enumerate(ozet.get_children(), start=2):
        values = ozet.item(item, "values")
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Set cell size
            ws.column_dimensions[get_column_letter(col_idx)].width = max(ws.column_dimensions[get_column_letter(col_idx)].width, len(str(value)) + 2)
    # Add Profit title and total amount
    net_kar_column = len(columns) + 1
    ws.cell(row=1, column=net_kar_column, value="Net Kâr").fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Mavi renk
    ws.cell(row=2, column=net_kar_column, value=text.get()).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Yeşil renk
    # Set cell size
    ws.column_dimensions[get_column_letter(net_kar_column)].width = max(len("Net Kâr") + 2, 10)  # Minimum 10 karakter genişlik
    # Save the Excel file
    wb.save(filename)


def save_file():# Function to save file
    # Requesting the file path and file name from the user
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if filename:
        export_to_excel(ozet, filename)


excel_button = ctk.CTkButton(window, text="Excel'e Aktar", fg_color="blue", hover_color="dark blue", font=button_font,command=save_file)
excel_button.grid(column=4, row=4, sticky="ne",padx=15)


def kar_hesapla(): # Function to calculate profit
    # Create a database connection
    baglanti = sql.connect("Stok.sqlite")
    scursor = baglanti.cursor()
    ucursor = baglanti.cursor()

    # Toplam karı hesaplamak için bir değişken tanımla
    toplam_kar = 0

    # Retrieve all sales data from the Sale table
    scursor.execute("SELECT UrunKodu, Miktar, Tutar FROM Satis")
    satis_veri = scursor.fetchall()

    # Calculate profit for each line and add to total profit
    for satir in satis_veri:
        kod, miktar, tutar = satir

        # Get product cost from "Product" table
        ucursor.execute("SELECT UrunMaliyet FROM Urun WHERE UrunKodu=?", (kod,))
        urun_veri = ucursor.fetchone()

        if urun_veri:  # If product data exist
            urun_maliyeti = urun_veri[0]
            # Quantity and product cost are converted into integers and calculated as profit.
            kar = tutar - (int(miktar) * int(urun_maliyeti))
            toplam_kar += kar  # Add profit to total profit

    # Close database connection
    baglanti.close()
    text.delete(0, tk.END)
    text.insert(tk.END, toplam_kar)
    text.after(1000,kar_hesapla)


kar_hesapla()


def fetch_data(): # Function to retrieve data from database
    baglanti = sql.connect("Stok.sqlite")
    cursor = baglanti.cursor()
    cursor.execute("SELECT * FROM Satis")
    data = cursor.fetchall()
    return data


# TreeView update function
def update_treeview():
    # Get data
    data = fetch_data()
    # Clear existing data
    ozet.delete(*ozet.get_children())
    # Add new data to TreeView
    for row in data:
        ozet.insert("", "end", values=row)
    # Call the auto-update function again
    ozet.after(50, update_treeview)  # Call again after 0.05 seconds


update_treeview()


def verileri_getir():
    baglanti = sql.connect("Stok.sqlite")
    cursor = baglanti.cursor()
    cursor.execute("SELECT * FROM Satis")
    veriler = cursor.fetchall()
    baglanti.close()
    return veriler


def verileri_goster():
    veriler = verileri_getir()
    for row in veriler:
        ozet.insert("", "end", values=row)


verileri_getir()
verileri_goster()


def create_muskayit():
    global muskayits
    muskayits = MusKayit()


muskayit = ctk.CTkButton(window,
                         text="Müşteri Kayıt",
                         fg_color="blue",
                         hover_color="dark blue",
                         font=button_font,
                         command=create_muskayit)


class MusKayit(ctk.CTkToplevel):
    def __init__(self):
        super().__init__()
        self.title("Müşteri Kayıt")
        self.geometry("400x600")
        self.columnconfigure((0, 1), uniform="a", weight=1)
        self.rowconfigure((0, 1, 2, 3, 4, 5), uniform="a", weight=1)
        self.attributes("-topmost", True)
        self.resizable(False,False)
        ctk.CTkLabel(self, text="Adı:").grid(column=0, row=0, sticky='swe')
        ctk.CTkLabel(self, text="Soyadı:").grid(column=0, row=1, sticky='swe')
        ctk.CTkLabel(self, text="Telefon:").grid(column=0, row=2, sticky='swe')
        ctk.CTkLabel(self, text="E-posta:").grid(column=0, row=3, sticky='swe')
        ctk.CTkLabel(self, text="Adres:").grid(column=0, row=4, sticky='swe')
        self.e1 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e2 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e3 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e4 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e5 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e1.grid(column=2, row=0, sticky='se')
        self.e2.grid(column=2, row=1, sticky='se')
        self.e3.grid(column=2, row=2, sticky='se')
        self.e4.grid(column=2, row=3, sticky='se')
        self.e5.grid(column=2, row=4, sticky='se')

        def kaydet():
            vt = sql.connect("Stok.sqlite")
            imlec = vt.cursor()
            adi = self.e1.get()
            soyadi = self.e2.get()
            tel = self.e3.get()
            eposta = self.e4.get()
            adres = self.e5.get()
            imlec.execute("INSERT INTO Musteri (Adi,Soyadi,Telefon,E_posta,Adres) VALUES(?,?,?,?,?)",
                          (adi, soyadi, tel, eposta, adres))
            self.e1.delete(0,100)
            self.e2.delete(0,100)
            self.e3.delete(0,100)
            self.e4.delete(0,100)
            self.e5.delete(0,100)
            vt.commit()
            vt.close()
        ctk.CTkButton(self, text="Kaydet", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=kaydet).grid(column=1, row=5)


def create_muslist():
    global muslists
    muslists = MusList()


muslist = ctk.CTkButton(window,
                        text="Müşteri Listesi",
                        fg_color="blue",
                        hover_color="dark blue",
                        font=button_font,
                        command=create_muslist)


class MusList(ctk.CTkToplevel):
    def __init__(self):
        super().__init__()
        self.tv = ttk.Treeview(self, columns=("Adı", "Soyadı", "Telefon", "E-posta", "Adres"), show="headings")
        self.title("Müşteri Listesi")
        self.geometry("800x500")
        self.columnconfigure((0, 1), uniform="a", weight=1)
        self.columnconfigure(2, weight=5, uniform="a")
        self.rowconfigure((0, 1, 2, 3, 4, 5), uniform="a", weight=1)
        self.attributes("-topmost", True)
        self.ayarla_sutun_genislikleri()
        ctk.CTkLabel(self, text="Adı:").grid(column=0, row=0, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Soyadı:").grid(column=0, row=1, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Telefon:").grid(column=0, row=2, sticky='w', padx=5)
        ctk.CTkLabel(self, text="E-posta:").grid(column=0, row=3, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Adres:").grid(column=0, row=4, sticky='w', padx=5)
        self.e1 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e1.grid(column=1, row=0, sticky='w')
        self.e2 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e2.grid(column=1, row=1, sticky='w')
        self.e3 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e3.grid(column=1, row=2, sticky='w')
        self.e4 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e4.grid(column=1, row=3, sticky='w')
        self.e5 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e5.grid(column=1, row=4, sticky='w')
        self.tv.heading("Adı",text="Adı",anchor="sw")
        self.tv.heading("Soyadı", text="Soyadı",anchor="sw")
        self.tv.heading("Telefon", text="Telefon",anchor="sw")
        self.tv.heading("E-posta", text="E-posta",anchor="sw")
        self.tv.heading("Adres", text="Adres",anchor="sw")
        self.tv.grid(column=2, row=0, sticky="nsew", padx=20, pady=20, rowspan=6)
        self.tv.bind("<Double-1>", self.sec) #TreeView double click olayında gerçekleşecek fonksiyonu belirler.
        ctk.CTkButton(self, text="Güncelle", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=self.guncelle).grid(column=0, row=5, padx=5)
        ctk.CTkButton(self, text="Sil", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=self.sil).grid(column=1, row=5, padx=5)
        self.verileri_goster()

    def ayarla_sutun_genislikleri(self):
        for column in self.tv["columns"]:
            self.tv.column(column, width=ctk.CTkFont().measure(column))  # Sütun başlıklarının genişliği
            for cell in self.tv.get_children():
                cell_width = ctk.CTkFont().measure(self.tv.item(cell, "values")[self.tv["columns"].index(column)])
                if cell_width > self.tv.column(column, width=None):
                    self.tv.column(column, width=cell_width)  # Hücre içeriğinin genişliği
    # Adjusting column widths after creating a treeview

    def sec(self, event):
        item = self.tv.selection()[0]
        secilen = self.tv.item(item, "values")
        self.e1.delete(0, tk.END)
        self.e1.insert(0, secilen[0])
        self.e2.delete(0, tk.END)
        self.e2.insert(0, secilen[1])
        self.e3.delete(0, tk.END)
        self.e3.insert(0, secilen[2])
        self.e4.delete(0, tk.END)
        self.e4.insert(0, secilen[3])
        self.e5.delete(0, tk.END)
        self.e5.insert(0, secilen[4])
        # It cleans the inside of the entries and transfers the selected data to the entries.

    def guncelle(self):
        # Get the ID of the selected item
        secili_item = self.tv.selection()[0]
        # Get data of selected item
        secilen_veri = self.tv.item(secili_item, "values")
        # Get new data
        yeni_veri = (
            self.e1.get(),
            self.e2.get(),
            self.e3.get(),
            self.e4.get(),
            self.e5.get()
        )
        # Update the database
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("UPDATE Musteri SET Adi=?, Soyadi=?, Telefon=?, 'E_posta'=?, Adres=? WHERE Telefon=?",
                       (*yeni_veri, secilen_veri[2]))
        baglanti.commit()
        baglanti.close()
        # Update item in treeview
        self.tv.item(secili_item, values=yeni_veri)

    def sil(self):
        # Get the ID of the selected item
        secili_item = self.tv.selection()[0]
        # Get data of selected item
        secilen_veri = self.tv.item(secili_item, "values")
        # Delete record from database
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("DELETE FROM Musteri WHERE Telefon=?", (secilen_veri[2],))
        baglanti.commit()
        baglanti.close()
        # Delete item from treeview
        self.tv.delete(secili_item)

    def verileri_getir(self):
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("SELECT * FROM Musteri")
        veriler = cursor.fetchall()
        baglanti.close()
        return veriler

    def verileri_goster(self):
        veriler = self.verileri_getir()
        for row in veriler:
            self.tv.insert("", "end", values=row)


def create_urunkayit():
    global urunkayits
    urunkayits = UrunKayit()


class UrunKayit(ctk.CTkToplevel):
    def __init__(self):
        super().__init__()
        self.title("Ürün Kayıt")
        self.geometry("400x600")
        self.columnconfigure((0, 1), uniform="a", weight=1)
        self.rowconfigure((0, 1, 2, 3, 4, 5), uniform="a", weight=1)
        self.resizable(False, False)
        self.attributes("-topmost", True)
        ctk.CTkLabel(self, text="Ürün Adı:").grid(column=0, row=0, sticky='swe')
        ctk.CTkLabel(self, text="Ürün Kodu:").grid(column=0, row=1, sticky='swe')
        ctk.CTkLabel(self, text="Ürün Sayısı:").grid(column=0, row=2, sticky='swe')
        ctk.CTkLabel(self, text="Birim Maliyet:").grid(column=0, row=3, sticky='swe')
        ctk.CTkLabel(self, text="Birim Satış:").grid(column=0, row=4, sticky='swe')
        e1 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        e1.grid(column=2, row=0, sticky='se')
        e2 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        e2.grid(column=2, row=1, sticky='se')
        e3 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        e3.grid(column=2, row=2, sticky='se')
        e4 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        e4.grid(column=2, row=3, sticky='se')
        e5 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        e5.grid(column=2, row=4, sticky='se')

        def kaydet():
            vt = sql.connect("Stok.sqlite")
            imlec = vt.cursor()
            adi = e1.get()
            soyadi = e2.get()
            tel = e3.get()
            eposta = e4.get()
            adres = e5.get()
            imlec.execute("INSERT INTO Urun (UrunAdi,UrunKodu,UrunSayisi,UrunMaliyet,UrunSatis) VALUES(?,?,?,?,?)",(adi, soyadi, tel, eposta, adres))
            e1.delete(0,100)
            e2.delete(0,100)
            e3.delete(0,100)
            e4.delete(0,100)
            e5.delete(0,100)
            vt.commit()
            vt.close()
        ctk.CTkButton(self, text="Kaydet", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=kaydet).grid(column=1, row=5)


urunkayit = ctk.CTkButton(window,
                          text="Ürün Kayıt",
                          fg_color="blue",
                          hover_color="dark blue",
                          font=button_font,
                          command=create_urunkayit)


def create_urunlist():
    global urunlists
    urunlists = UrunList()


urunlist = ctk.CTkButton(window,
                         text="Ürün Listesi",
                         fg_color="blue",
                         hover_color="dark blue",
                         font=button_font,
                         command=create_urunlist)


class UrunList(ctk.CTkToplevel):
    def __init__(self):
        super().__init__()
        self.tv = ttk.Treeview(self, columns=("Ürün Adı", "Ürün Kodu", "Miktar", "Birim Maliyet", "Birim Satış"), show="headings")
        self.title("Ürün Listesi")
        self.geometry("800x500")
        self.columnconfigure((0, 1), uniform="a", weight=1)
        self.columnconfigure(2, weight=5, uniform="a")
        self.rowconfigure((0, 1, 2, 3, 4, 5), uniform="a", weight=1)
        self.attributes("-topmost", True)
        self.ayarla_sutun_genislikleri()
        ctk.CTkLabel(self, text="Ürün Adı:").grid(column=0, row=0, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Ürün Kodu:").grid(column=0, row=1, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Miktar:").grid(column=0, row=2, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Birim Maliyet:").grid(column=0, row=3, sticky='w', padx=5)
        ctk.CTkLabel(self, text="Birim Satış:").grid(column=0, row=4, sticky='w', padx=5)
        self.e1 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e1.grid(column=1, row=0, sticky='w')
        self.e2 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e2.grid(column=1, row=1, sticky='w')
        self.e3 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e3.grid(column=1, row=2, sticky='w')
        self.e4 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e4.grid(column=1, row=3, sticky='w')
        self.e5 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e5.grid(column=1, row=4, sticky='w')
        self.tv.heading("Ürün Adı", text="Ürün Adı", anchor="sw")
        self.tv.heading("Ürün Kodu", text="Ürün Kodu", anchor="sw")
        self.tv.heading("Miktar", text="Miktar", anchor="sw")
        self.tv.heading("Birim Maliyet", text="Birim Maliyet", anchor="sw")
        self.tv.heading("Birim Satış", text="Birim Satış", anchor="sw")
        self.tv.grid(column=2, row=0, sticky="nsew", padx=20, pady=20, rowspan=6)
        self.tv.bind("<Double-1>", self.sec)
        ctk.CTkButton(self, text="Güncelle", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=self.guncelle).grid(column=0, row=5, padx=5)
        ctk.CTkButton(self, text="Sil", font=button_font,
                      fg_color="blue", hover_color="dark blue", command=self.sil).grid(column=1, row=5, padx=5)
        self.verileri_goster()

    def ayarla_sutun_genislikleri(self):
        for column in self.tv["columns"]:
            self.tv.column(column, width=ctk.CTkFont().measure(column))  # Sütun başlıklarının genişliği
            for cell in self.tv.get_children():
                cell_width = ctk.CTkFont().measure(self.tv.item(cell, "values")[self.tv["columns"].index(column)])
                if cell_width > self.tv.column(column, width=None):
                    self.tv.column(column, width=cell_width)  # Width of cell content
# Adjusting column widths after creating a treeview

    def sec(self, event):
        item = self.tv.selection()[0]
        secilen = self.tv.item(item, "values")
        self.e1.delete(0, tk.END)
        self.e1.insert(0, secilen[0])
        self.e2.delete(0, tk.END)
        self.e2.insert(0, secilen[1])
        self.e3.delete(0, tk.END)
        self.e3.insert(0, secilen[2])
        self.e4.delete(0, tk.END)
        self.e4.insert(0, secilen[3])
        self.e5.delete(0, tk.END)
        self.e5.insert(0, secilen[4])

    def guncelle(self):
        # Get the ID of the selected item
        secili_item = self.tv.selection()[0]
        # Get data of selected item
        secilen_veri = self.tv.item(secili_item, "values")
        # Get new data
        yeni_veri = (
            self.e1.get(),
            self.e2.get(),
            self.e3.get(),
            self.e4.get(),
            self.e5.get()
        )
        # Update the database
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("UPDATE Urun SET UrunAdi=?,UrunKodu=?,UrunSayisi=?,UrunMaliyet=?,UrunSatis=? WHERE UrunKodu=?",
                       (*yeni_veri, secilen_veri[1]))
        baglanti.commit()
        baglanti.close()
        # Update item in treeview
        self.tv.item(secili_item, values=yeni_veri)

    def sil(self):
        # Get the ID of the selected item
        secili_item = self.tv.selection()[0]
        # Get data of selected item
        secilen_veri = self.tv.item(secili_item, "values")
        # Delete record from database
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("DELETE FROM Urun WHERE UrunKodu=?", (secilen_veri[1],))
        baglanti.commit()
        baglanti.close()
        # Delete item from treeview
        self.tv.delete(secili_item)

    def verileri_getir(self):
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("SELECT * FROM Urun")
        veriler = cursor.fetchall()
        baglanti.close()
        return veriler

    def verileri_goster(self):
        veriler = self.verileri_getir()
        for row in veriler:
            self.tv.insert("", "end", values=row)


def create_satis():
    global satis
    satis = Satis()


satis = ctk.CTkButton(window, text="Satış", fg_color="blue", hover_color="dark blue", font=button_font, command=create_satis)


class Satis(ctk.CTkToplevel):
    def __init__(self):
        super().__init__()
        self.title("Satış")
        self.geometry("900x600")
        self.attributes("-topmost", True)
        self.columnconfigure((0,1,2,3,4), weight=1, uniform="a")
        self.rowconfigure((0,1,2,3,4,5,6), weight=1, uniform="a")
        self.tv1 = ttk.Treeview(self, columns=("Adı", "Soyadı","Telefon"), show="headings")
        self.tv2 = ttk.Treeview(self, columns=("Ürün Kodu", "Ürün Adı", "Miktar", "Birim Fiyat"), show="headings")
        self.tv1.heading("Adı", text="Adı", anchor="sw")
        self.tv1.heading("Soyadı", text="Soyadı", anchor="sw")
        self.tv1.heading("Telefon", text="Telefon", anchor="sw")
        self.tv2.heading("Ürün Kodu", text="Ürün Kodu", anchor="sw")
        self.tv2.heading("Ürün Adı", text="Ürün Adı", anchor="sw")
        self.tv2.heading("Miktar", text="Miktar", anchor="sw")
        self.tv2.heading("Birim Fiyat", text="Birim Fiyat", anchor="sw")
        self.tv1.grid(column=2,columnspan=2,row=0,rowspan=3,sticky="nsew", padx=30, pady=30)
        self.tv2.grid(column=2,columnspan=2,row=3,rowspan=3, sticky="nsew", padx=30, pady=30)
        self.tv1.bind("<Double-1>", self.mus_sec)
        self.tv2.bind("<Double-1>", self.urun_sec)
        self.verileri_goster()
        self.l1 = ctk.CTkLabel(self, text="Telefon:")
        self.l1.grid(column=0, row=0, sticky="w", padx=20)
        self.l2 = ctk.CTkLabel(self, text="Adı:")
        self.l2.grid(column=0, row=1, sticky="w", padx=20)
        self.l3 = ctk.CTkLabel(self, text="Soyadı:")
        self.l3.grid(column=0, row=2, sticky="w", padx=20)
        self.l4 = ctk.CTkLabel(self, text="Ürün Kodu:")
        self.l4.grid(column=0, row=3, sticky="w", padx=20)
        self.l5 = ctk.CTkLabel(self, text="Ürün Adı:")
        self.l5.grid(column=0, row=4, sticky="w", padx=20)
        self.l6 = ctk.CTkLabel(self, text="Birim Fiyat:")
        self.l6.grid(column=0, row=5, sticky="w", padx=20)
        self.l7 = ctk.CTkLabel(self, text="Miktar:")
        self.l7.grid(column=0, row=6, sticky="w", padx=20)
        self.l8 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.l8.grid(column=1, row=6, sticky="w")
        self.e1 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue",)
        self.e1.grid(column=1, row=0, sticky="nsew",pady=30)
        self.e2 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e2.grid(column=1, row=1, sticky="nsew",pady=30)
        self.e3 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e3.grid(column=1, row=2, sticky="nsew",pady=30)
        self.e4 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e4.grid(column=1, row=4, sticky="nsew",pady=30)
        self.e5 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e5.grid(column=1, row=5, sticky="nsew",pady=30)
        self.e6 = ctk.CTkEntry(self, fg_color="White", text_color="Black", border_color="Blue")
        self.e6.grid(column=1, row=3, sticky="nsew",pady=30)
        self.b1 = ctk.CTkButton(self, text="Tamamla", font=button_font,fg_color="blue", hover_color="dark blue", command=self.tamamla)
        self.b1.grid(column=2, row=6, padx=5)

    def mus_sec(self, event):
        item = self.tv1.selection()[0]
        secilen = self.tv1.item(item, "values")
        self.e1.delete(0, tk.END)
        self.e1.insert(0, secilen[2])
        self.e2.delete(0, tk.END)
        self.e2.insert(0, secilen[0])
        self.e3.delete(0, tk.END)
        self.e3.insert(0, secilen[1])

    def urun_sec(self, event):
        item = self.tv2.selection()[0]
        secilen = self.tv2.item(item, "values")
        self.e6.delete(0, tk.END)
        self.e6.insert(0, secilen[0])
        self.e4.delete(0, tk.END)
        self.e4.insert(0, secilen[1])
        self.e5.delete(0, tk.END)
        self.e5.insert(0, secilen[3])

    def verileri_getir(self):
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("SELECT Adi, Soyadi, Telefon FROM Musteri")
        veriler = cursor.fetchall()
        baglanti.close()
        return veriler

    def urun_getir(self):
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        cursor.execute("SELECT UrunKodu, UrunAdi, UrunSayisi, UrunSatis FROM Urun")
        urunler = cursor.fetchall()
        baglanti.close()
        return urunler

    def verileri_goster(self):
        veriler = self.verileri_getir()
        urunler = self.urun_getir()
        for row in veriler:
            self.tv1.insert("", "end", values=row)
        for row in urunler:
            self.tv2.insert("","end", values=row)

    def tamamla(self):
        baglanti = sql.connect("Stok.sqlite")
        cursor = baglanti.cursor()
        tel = self.e1.get()
        adi = self.e2.get()
        soyadi = self.e3.get()
        kod = self.e6.get()
        u_adi = self.e4.get()
        miktar = self.l8.get()
        tutar = int(self.l8.get()) * int(self.e5.get())
        cursor.execute("INSERT INTO Satis (Adi,Soyadi,Telefon,UrunAdi,UrunKodu,Miktar,Tutar) VALUES(?,?,?,?,?,?,?)",(adi, soyadi, tel, u_adi, kod, miktar, tutar))
        cursor.execute("SELECT UrunSayisi FROM Urun WHERE UrunKodu=?",(self.e6.get()))
        stok = cursor.fetchone()
        ystok = int(stok[0]) - int(self.l8.get())
        cursor.execute("UPDATE Urun SET UrunSayisi=? WHERE UrunKodu=?",(ystok,self.e6.get()))
        baglanti.commit()
        baglanti.close()
        self.e1.delete(0,100)
        self.e2.delete(0,100)
        self.e3.delete(0,100)
        self.e4.delete(0,100)
        self.e5.delete(0,100)
        self.e6.delete(0,100)
        self.l8.delete(0,100)


muskayit.grid(row=0, column=0, sticky="w", padx=10)
muslist.grid(row=1, column=0, sticky="w", padx=10)
urunkayit.grid(row=2, column=0, sticky="w", padx=10)
urunlist.grid(row=3, column=0, sticky="w", padx=10)
satis.grid(row=4, column=0, sticky="w", padx=10)
window.mainloop()
